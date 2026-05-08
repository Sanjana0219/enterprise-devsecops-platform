from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Security Scan Report"

headers = [
    "Scanner",
    "Critical",
    "High",
    "Medium",
    "Low",
    "Status"
]

ws.append(headers)

# Example enterprise vulnerability summary
data = [
    ["Snyk", 1, 6, 4, 0, "FAIL"],
    ["Trivy", 0, 3, 2, 1, "WARN"],
    ["Checkov", 0, 8, 2, 0, "FAIL"],
    ["Gitleaks", 2, 0, 0, 0, "FAIL"],
    ["OWASP Dependency Check", 1, 4, 5, 2, "FAIL"],
    ["OPA Policy Gate", 0, 2, 1, 0, "FAIL"]
]

for row in data:
    ws.append(row)

# Header styling
header_fill = PatternFill(
    start_color="000000",
    end_color="000000",
    fill_type="solid"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Severity colors
red_fill = PatternFill(
    start_color="FF4C4C",
    end_color="FF4C4C",
    fill_type="solid"
)

orange_fill = PatternFill(
    start_color="FFA500",
    end_color="FFA500",
    fill_type="solid"
)

yellow_fill = PatternFill(
    start_color="FFFF99",
    end_color="FFFF99",
    fill_type="solid"
)

green_fill = PatternFill(
    start_color="90EE90",
    end_color="90EE90",
    fill_type="solid"
)

# Apply row styling
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):

    critical = row[1].value
    high = row[2].value
    medium = row[3].value
    status = row[5].value

    if critical > 0:
        for cell in row:
            cell.fill = red_fill

    elif high > 0:
        for cell in row:
            cell.fill = orange_fill

    elif medium > 0:
        for cell in row:
            cell.fill = yellow_fill

    elif status == "PASS":
        for cell in row:
            cell.fill = green_fill

# Adjust column widths
column_widths = {
    "A": 30,
    "B": 12,
    "C": 12,
    "D": 12,
    "E": 12,
    "F": 15
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save workbook
wb.save("security-report.xlsx")

print("Enterprise Excel security report generated successfully.")